from django.shortcuts import render, redirect
from openpyxl import load_workbook
from .models import JobDescription


def file_upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        files = request.FILES.getlist('myfile')
        processed_data = []

        for each_file in files:
            file_name = each_file.name
            work_book = load_workbook(each_file)
            sheet = work_book.worksheets[0]
            total_rows_count = sheet.max_row
            rows_ignored = 1
            rows_processed = 0

            for row_count in range(2, total_rows_count):
                if total_rows_count <= 10000:
                    for row_cells in sheet.iter_rows(min_row=row_count,
                                                     max_row=row_count):
                        row_data = []
                        for cell in row_cells:
                            row_data.append(cell.value)

                        if JobDescription.objects.filter(
                                job_id=row_data[0]).exists():
                            rows_ignored += 1
                        else:
                            rows_processed += 1
                            JobDescription.objects.create(job_id=row_data[0],
                                                          role=row_data[1],
                                                          level=row_data[2],
                                                          primary_skills=row_data[3],
                                                          secondary_skills=row_data[4])
                else:
                    break

            file_context = {
                "file_name": file_name,
                "total_rows_count": total_rows_count,
                "rows_ignored": rows_ignored,
                "rows_processed": rows_processed,
                "Jobs_description_created": rows_processed
            }

            processed_data.append(file_context)

        request.session['data'] = processed_data

        return redirect('dashboard')
    else:
        return render(request, 'file_upload.html')


def dashboard(request):
    data = request.session.get('data')
    return render(request, 'dashboard.html', {'data': data})
